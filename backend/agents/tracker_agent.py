"""
Tracker Agent
-------------
Manages the application pipeline and generates exports.
"""

from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STATUS_COLORS = {
    "saved":       "E2E8F0",
    "applied":     "DBEAFE",
    "shortlisted": "FEF3C7",
    "interview":   "D1FAE5",
    "offer":       "A7F3D0",
    "rejected":    "FEE2E2",
    "ghosted":     "F3F4F6",
    "withdrawn":   "F9FAFB",
}


def generate_tracker_excel(applications: list[dict]) -> bytes:
    """
    Generate a professional Excel tracker for all applications.
    Returns bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = [
        ("Job Title", 28),
        ("Company", 22),
        ("Location", 18),
        ("Job Type", 14),
        ("Source", 14),
        ("Status", 14),
        ("Applied Date", 16),
        ("Salary Range", 16),
        ("Match Score", 13),
        ("Outreach Sent", 14),
        ("Outreach Status", 16),
        ("Notes", 30),
        ("Apply Link", 35),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, app in enumerate(applications, 2):
        status = app.get("status", "applied")
        row_fill_color = STATUS_COLORS.get(status, "FFFFFF")
        row_fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type="solid")

        salary_min = app.get("salary_min")
        salary_max = app.get("salary_max")
        currency = app.get("salary_currency", "")
        if salary_min and salary_max:
            salary_str = f"{currency} {salary_min:,} - {salary_max:,}"
        elif salary_min:
            salary_str = f"{currency} {salary_min:,}+"
        else:
            salary_str = "Not disclosed"

        applied_date = app.get("applied_at")
        if applied_date:
            try:
                applied_date = datetime.fromisoformat(applied_date).strftime("%d %b %Y")
            except Exception:
                applied_date = str(applied_date)

        match_score = app.get("match_score")
        match_str = f"{match_score:.0f}%" if match_score else "—"

        outreach = app.get("outreach_sent")
        outreach_sent = "Yes" if outreach else "No"
        outreach_status = outreach[0].get("status", "—") if outreach else "—"

        row_data = [
            app.get("job_title", ""),
            app.get("company", ""),
            app.get("location", ""),
            app.get("job_type", "").replace("_", " ").title(),
            app.get("source", "").replace("_", " ").title(),
            status.replace("_", " ").title(),
            applied_date or "—",
            salary_str,
            match_str,
            outreach_sent,
            outreach_status.title(),
            app.get("notes", ""),
            app.get("source_url", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [1, 12]))

            # Hyperlink for apply URL
            if col_idx == 13 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="4F46E5", underline="single")

        ws.row_dimensions[row_idx].height = 22

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    status_counts = {}
    for app in applications:
        s = app.get("status", "applied")
        status_counts[s] = status_counts.get(s, 0) + 1

    ws2.cell(1, 1, "Status").font = Font(bold=True)
    ws2.cell(1, 2, "Count").font = Font(bold=True)
    for i, (status, count) in enumerate(status_counts.items(), 2):
        ws2.cell(i, 1, status.replace("_", " ").title())
        ws2.cell(i, 2, count)

    ws2.cell(len(status_counts) + 3, 1, "Total Applications").font = Font(bold=True)
    ws2.cell(len(status_counts) + 3, 2, len(applications)).font = Font(bold=True)

    response_rate = sum(1 for a in applications if a.get("status") in ["shortlisted", "interview", "offer"]) / max(len(applications), 1) * 100
    ws2.cell(len(status_counts) + 4, 1, "Response Rate")
    ws2.cell(len(status_counts) + 4, 2, f"{response_rate:.1f}%")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 10

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
